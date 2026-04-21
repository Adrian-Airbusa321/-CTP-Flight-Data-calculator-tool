#!/usr/bin/env python3
"""Transatlantic Fuel & Flight-Time Calculator.

Route distance data courtesy of @ianolaf — thanks for compiling
the transatlantic city-pair distance matrix used by this tool.

Reads:
  ../Data/Aircraft/Aircraft Performance Data.xlsx
  ../Data/Route/Pairs and Distances.xlsx

Computes phase-based fuel burn and block time for a selected route + aircraft.

Usage:
  python fuel_calculator.py                           # interactive
  python fuel_calculator.py --list-aircraft
  python fuel_calculator.py --list-routes
  python fuel_calculator.py --origin EGLL --dest KJFK --aircraft B787-9
  python fuel_calculator.py --origin EGLL --dest KJFK --aircraft B787-9 \\
                            --payload 25000 --wind 40

Model (simplified airline flight-planning):
  Block fuel  = Taxi-out + Climb + Cruise + Descent + Taxi-in
  Total fuel  = Block + Contingency (5 % trip) + Alternate (200 nm default)
                     + Final reserve (30 min holding)
  Block time  = sum of phase times
  TAS         = Mach × 573.5 kt   (ISA above tropopause)
  Climb burn  ≈ 1.6 × cruise burn   Descent burn ≈ 0.35 × cruise burn
  Final burn  ≈ 0.70 × cruise burn   Alternate trip with 20 % climb/app overhead
  Taxi       : narrow-body 10 kg/min, wide-body twin 20, quad 35

All fuel figures are kg, distances nm, speeds kt.  Expect operational
variation of ±5-10 % vs real-world flight plans.
"""
from __future__ import annotations

import argparse
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Optional, Tuple

try:
    from openpyxl import load_workbook
except ImportError:
    print("This tool needs openpyxl.  Install with:  pip install openpyxl",
          file=sys.stderr)
    sys.exit(2)

# Make sure Unicode characters in the report (→, ─, ═, ✓) don't crash the
# Windows console when its default code page is cp1252. We neither want to crash the sim during CTP! xD
for _stream in (sys.stdout, sys.stderr):
    try:
        _stream.reconfigure(encoding="utf-8", errors="replace")  # type: ignore[attr-defined]
    except Exception:
        pass


# Paths (RIP Pathfinder)

TOOL_DIR = Path(__file__).resolve().parent
BASE_DIR = TOOL_DIR.parent

AIRCRAFT_CANDIDATES = [
    BASE_DIR / "Data" / "Aircraft" / "Aircraft Performance Data.xlsx",
    BASE_DIR / "Data" / "Aircraft" / "Aircraft Performance Data (redesigned).xlsx",
]
ROUTES_CANDIDATES = [
    BASE_DIR / "Data" / "Route" / "Pairs and Distances.xlsx",
]


def _pick_existing(candidates: List[Path]) -> Optional[Path]:
    """Return the first existing, readable workbook from the list."""
    import zipfile
    for p in candidates:
        if not p.exists():
            continue
        try:
            with zipfile.ZipFile(p, "r"):
                return p
        except (zipfile.BadZipFile, PermissionError):
            continue
    return None



# Physics / planning constants (I like physics, hopefully I qualify for the national Olympiad next year)

TAS_PER_MACH_KT          = 573.5      # ISA speed of sound above tropopause (kt)
MIN_PER_H                = 60.0

CLIMB_DIST_NM_NARROW     = 130        # average ground distance covered in climb
CLIMB_DIST_NM_WIDE       = 160
DESCENT_DIST_NM          = 130

CLIMB_TIME_MIN_NARROW    = 22
CLIMB_TIME_MIN_WIDE      = 27
DESCENT_TIME_MIN         = 22

CLIMB_BURN_FACTOR        = 1.60       # *cruise burn (Fun Fact: Can't appy this to rockets, since they don't burn fuel while "cruising" and even 500000*0=0)
DESCENT_BURN_FACTOR      = 0.35

TAXI_OUT_MIN_DEFAULT     = 15
TAXI_IN_MIN_DEFAULT      = 7
TAXI_KGPMIN_NARROW       = 10
TAXI_KGPMIN_WIDE_TWIN    = 20
TAXI_KGPMIN_QUAD         = 35

CONTINGENCY_PCT          = 0.05
ALTERNATE_DIST_NM        = 200
ALTERNATE_OVERHEAD       = 1.20       # add climb + approach
FINAL_RESERVE_MIN        = 30
FINAL_RESERVE_BURN_FCT   = 0.70

PAX_WEIGHT_KG            = 100        # pax + bag + cargo share
LOAD_FACTOR_PAX          = 0.85
LOAD_FACTOR_CARGO        = 0.70



# Data classes 

@dataclass
class Aircraft:
    id:              str
    manufacturer:    str
    family:          str
    variant:         str
    engines:         str
    mtow:            float
    oew:             Optional[float]
    mzfw:            Optional[float]
    mlw:             Optional[float]
    max_fuel:        float
    max_range:       float
    cruise_mach:     float
    mmo:             Optional[float]
    typical_fl:      int
    ceiling:         Optional[int]
    max_pax_or_cargo_t: Optional[float]
    cruise_burn:     float            # kg/h
    tsfc_note:       Optional[str]
    source:          Optional[str]

    # --- derived ---
    @property
    def is_freighter(self) -> bool:
        v = self.variant.upper().strip()
        return v.endswith("F") or "FREIGHTER" in v

    @property
    def engine_count(self) -> int:
        if self.family in ("747", "747-8", "A340"):
            return 4
        if "(4x)" in (self.engines or "").lower():
            return 4
        return 2

    @property
    def is_wide_body(self) -> bool:
        return self.family not in ("A320ceo", "A320neo", "737NG", "737MAX")

    @property
    def taxi_burn_kgpmin(self) -> float:
        if self.engine_count == 4:
            return TAXI_KGPMIN_QUAD
        return TAXI_KGPMIN_WIDE_TWIN if self.is_wide_body else TAXI_KGPMIN_NARROW


@dataclass
class Route:
    origin:       str
    destination:  str
    distance_nm:  float


@dataclass
class Calculation:
    aircraft:            Aircraft
    route:               Route
    # inputs
    payload_kg:          float
    wind_component_kt:   float
    load_factor:         float
    taxi_out_min:        float
    taxi_in_min:         float
    alternate_nm:        float
    # speeds
    cruise_tas_kt:       float
    cruise_gs_kt:        float
    # phase fuel
    taxi_out_kg:         float
    climb_kg:            float
    cruise_kg:           float
    descent_kg:          float
    taxi_in_kg:          float
    # phase time (h)
    taxi_out_h:          float
    climb_h:             float
    cruise_h:            float
    descent_h:           float
    taxi_in_h:           float
    # reserves
    contingency_kg:      float
    alternate_kg:        float
    final_reserve_kg:    float
    # totals
    trip_kg:             float
    block_kg:            float
    reserves_kg:         float
    total_fuel_kg:       float
    block_time_h:        float
    # weights
    zfw_kg:              Optional[float]
    tow_kg:              Optional[float]
    lw_kg:               Optional[float]
    # flags
    warnings:            List[str] = field(default_factory=list)
    # configurable reserves (kept at end so existing field order is unchanged)
    final_reserve_min:   float = FINAL_RESERVE_MIN
    contingency_pct:     float = CONTINGENCY_PCT



# Loading (Windows also always loading. POV: Me using Windows: "It wen't by like we were standing still" - It was the ME262)

def _is_icao(v) -> bool:
    return isinstance(v, str) and len(v) == 4 and v.isupper() and v.isalpha()


def _f(v) -> Optional[float]:
    try:
        return float(v) if v is not None and v != "" else None
    except (TypeError, ValueError):
        return None


def _i(v) -> Optional[int]:
    f = _f(v)
    return int(f) if f is not None else None


def load_aircraft(path: Path) -> Dict[str, Aircraft]:
    wb = load_workbook(path, data_only=True)
    ws = wb["Aircraft"] if "Aircraft" in wb.sheetnames else wb.active
    out: Dict[str, Aircraft] = {}
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row or not row[0]:
            continue
        ac_id = str(row[0]).strip()
        if not ac_id or ac_id.upper() == "ID":
            continue
        try:
            ac = Aircraft(
                id=ac_id,
                manufacturer=str(row[1] or "").strip(),
                family=str(row[2] or "").strip(),
                variant=str(row[3] or "").strip(),
                engines=str(row[4] or "").strip(),
                mtow=_f(row[5]) or 0.0,
                oew=_f(row[6]),
                mzfw=_f(row[7]),
                mlw=_f(row[8]),
                max_fuel=_f(row[9]) or 0.0,
                max_range=_f(row[10]) or 0.0,
                cruise_mach=_f(row[11]) or 0.80,
                mmo=_f(row[12]),
                typical_fl=_i(row[13]) or 350,
                ceiling=_i(row[14]),
                max_pax_or_cargo_t=_f(row[15]),
                cruise_burn=_f(row[16]) or 0.0,
                tsfc_note=(str(row[17]).strip() if row[17] else None),
                source=(str(row[18]).strip() if row[18] else None),
            )
        except Exception:
            continue
        if ac.mtow > 0 and ac.cruise_burn > 0:
            out[ac.id] = ac
    return out


def load_routes(path: Path) -> Dict[Tuple[str, str], float]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    # locate the header row – it's the one with several 4-letter ICAO codes
    header_idx = None
    for i, r in enumerate(rows):
        if sum(1 for v in r if _is_icao(v)) >= 5:
            header_idx = i
            break
    if header_idx is None:
        raise RuntimeError("Could not find ICAO header row in routes file.")

    header = rows[header_idx]
    dest_cols = [(idx, v) for idx, v in enumerate(header) if _is_icao(v)]

    routes: Dict[Tuple[str, str], float] = {}
    for r in rows[header_idx + 1:]:
        # the origin ICAO usually sits in column B or C
        origin = next((v for v in r[:4] if _is_icao(v)), None)
        if not origin:
            continue
        for col, dest in dest_cols:
            if col >= len(r):
                continue
            d = r[col]
            if isinstance(d, (int, float)) and d > 0:
                routes[(origin, dest)] = float(d)
                routes[(dest, origin)] = float(d)
    return routes



# Calculation (Erm actually, I should have gotten an one on my math exam. Mr. Lugisch is such a dork for giving me a two!)

def estimate_payload(ac: Aircraft, load_factor: Optional[float] = None) -> float:
    cap = ac.max_pax_or_cargo_t
    if cap is None:
        return 15000.0
    if ac.is_freighter:
        lf = load_factor if load_factor is not None else LOAD_FACTOR_CARGO
        return cap * 1000.0 * lf          # cap is tonnes
    lf = load_factor if load_factor is not None else LOAD_FACTOR_PAX
    return cap * PAX_WEIGHT_KG * lf       # cap is pax count


def calculate(
    ac:               Aircraft,
    route:            Route,
    payload_kg:       Optional[float] = None,
    wind_component_kt: float = 0.0,
    taxi_out_min:     float = TAXI_OUT_MIN_DEFAULT,
    taxi_in_min:      float = TAXI_IN_MIN_DEFAULT,
    alternate_nm:     float = ALTERNATE_DIST_NM,
    final_reserve_min: float = FINAL_RESERVE_MIN,
    contingency_pct:  float = CONTINGENCY_PCT,
) -> Calculation:
    if payload_kg is None:
        payload_kg = estimate_payload(ac)

    # Speeds
    tas = ac.cruise_mach * TAS_PER_MACH_KT
    gs  = max(tas - wind_component_kt, 50.0)

    # Climb
    if ac.is_wide_body:
        climb_dist  = CLIMB_DIST_NM_WIDE
        climb_time  = CLIMB_TIME_MIN_WIDE
    else:
        climb_dist  = CLIMB_DIST_NM_NARROW
        climb_time  = CLIMB_TIME_MIN_NARROW
    climb_h   = climb_time / MIN_PER_H
    climb_kg  = ac.cruise_burn * CLIMB_BURN_FACTOR * climb_h

    # Descent
    descent_h  = DESCENT_TIME_MIN / MIN_PER_H
    descent_kg = ac.cruise_burn * DESCENT_BURN_FACTOR * descent_h

    # Cruise (what's left of ground distance, flown at GS)
    cruise_dist = max(route.distance_nm - climb_dist - DESCENT_DIST_NM, 0.0)
    cruise_h    = cruise_dist / gs
    cruise_kg   = ac.cruise_burn * cruise_h

    # Taxi
    taxi_out_h  = taxi_out_min / MIN_PER_H
    taxi_in_h   = taxi_in_min  / MIN_PER_H
    taxi_out_kg = ac.taxi_burn_kgpmin * taxi_out_min
    taxi_in_kg  = ac.taxi_burn_kgpmin * taxi_in_min

    # Totals
    trip_kg      = climb_kg + cruise_kg + descent_kg
    block_kg     = taxi_out_kg + trip_kg + taxi_in_kg
    block_time_h = taxi_out_h + climb_h + cruise_h + descent_h + taxi_in_h

    # Reserves
    contingency   = trip_kg * contingency_pct
    alt_h         = alternate_nm / tas
    alternate_kg  = ac.cruise_burn * alt_h * ALTERNATE_OVERHEAD
    final_reserve = ac.cruise_burn * FINAL_RESERVE_BURN_FCT * (final_reserve_min / MIN_PER_H)
    reserves      = contingency + alternate_kg + final_reserve
    total_fuel    = block_kg + reserves

    # Weights
    if ac.oew is not None:
        zfw = ac.oew + payload_kg
        tow = zfw + trip_kg + reserves + taxi_in_kg     # taxi-out burned pre-TO
        lw  = tow - trip_kg
    else:
        zfw = tow = lw = None

    # Warnings (Last warning to @Nick Marinov, if he forgets to post the sloth nof the day again >:( )
    warn: List[str] = []
    if route.distance_nm > ac.max_range:
        warn.append(f"Route {route.distance_nm:,.0f} nm exceeds max range "
                    f"{ac.max_range:,.0f} nm")
    if total_fuel > ac.max_fuel:
        warn.append(f"Required fuel {total_fuel:,.0f} kg exceeds tank capacity "
                    f"{ac.max_fuel:,.0f} kg")
    if tow is not None and tow > ac.mtow:
        warn.append(f"Estimated TOW {tow:,.0f} kg exceeds MTOW {ac.mtow:,.0f} kg")
    if ac.mzfw is not None and zfw is not None and zfw > ac.mzfw:
        warn.append(f"ZFW {zfw:,.0f} kg exceeds MZFW {ac.mzfw:,.0f} kg")
    if ac.mlw is not None and lw is not None and lw > ac.mlw:
        warn.append(f"Landing weight {lw:,.0f} kg exceeds MLW {ac.mlw:,.0f} kg")

    lf = LOAD_FACTOR_CARGO if ac.is_freighter else LOAD_FACTOR_PAX
    return Calculation(
        aircraft=ac, route=route,
        payload_kg=payload_kg, wind_component_kt=wind_component_kt,
        load_factor=lf, taxi_out_min=taxi_out_min, taxi_in_min=taxi_in_min,
        alternate_nm=alternate_nm,
        cruise_tas_kt=tas, cruise_gs_kt=gs,
        taxi_out_kg=taxi_out_kg, climb_kg=climb_kg, cruise_kg=cruise_kg,
        descent_kg=descent_kg, taxi_in_kg=taxi_in_kg,
        taxi_out_h=taxi_out_h, climb_h=climb_h, cruise_h=cruise_h,
        descent_h=descent_h, taxi_in_h=taxi_in_h,
        contingency_kg=contingency, alternate_kg=alternate_kg,
        final_reserve_kg=final_reserve,
        trip_kg=trip_kg, block_kg=block_kg, reserves_kg=reserves,
        total_fuel_kg=total_fuel, block_time_h=block_time_h,
        zfw_kg=zfw, tow_kg=tow, lw_kg=lw, warnings=warn,
        final_reserve_min=final_reserve_min,
        contingency_pct=contingency_pct,
    )



# Formatting

def fmt_hm(h: float) -> str:
    total = int(round(h * 60))
    return f"{total // 60:2d}:{total % 60:02d}"


def fmt_kg(x: Optional[float], width: int = 9) -> str:
    if x is None:
        return f"{'—':>{width}}"
    return f"{x:>{width},.0f}".replace(",", " ")


def format_report(c: Calculation) -> str:
    ac, r = c.aircraft, c.route
    L = 64
    line = "─" * L
    dline = "═" * L
    out: List[str] = []
    out.append(dline)
    out.append("  TRANSATLANTIC FUEL & FLIGHT-TIME CALCULATOR")
    out.append(dline)
    out.append(f"  Route      :  {r.origin} → {r.destination}"
               f"      ({r.distance_nm:,.0f} nm)")
    out.append(f"  Aircraft   :  {ac.id}  —  {ac.manufacturer} {ac.variant}")
    out.append(f"  Engines    :  {ac.engines}")
    out.append(f"  Cruise     :  M{ac.cruise_mach:.3f} @ FL{ac.typical_fl:03d}   "
               f"TAS {c.cruise_tas_kt:,.0f} kt   GS {c.cruise_gs_kt:,.0f} kt")
    if c.wind_component_kt:
        sign = "head" if c.wind_component_kt > 0 else "tail"
        out.append(f"  Wind       :  {abs(c.wind_component_kt):,.0f} kt {sign}wind")
    kind = "cargo" if ac.is_freighter else "pax"
    out.append(f"  Payload    :  {fmt_kg(c.payload_kg)} kg   "
               f"({int(c.load_factor*100)} % {kind} load factor)")
    out.append(line)
    out.append("  PHASE                  TIME       DISTANCE       FUEL")
    out.append(line)
    cruise_dist = max(r.distance_nm - (CLIMB_DIST_NM_WIDE if ac.is_wide_body
                                       else CLIMB_DIST_NM_NARROW) - DESCENT_DIST_NM, 0)
    climb_dist_used = CLIMB_DIST_NM_WIDE if ac.is_wide_body else CLIMB_DIST_NM_NARROW
    out.append(f"   Taxi-out              {fmt_hm(c.taxi_out_h)}           —   {fmt_kg(c.taxi_out_kg)} kg")
    out.append(f"   Climb                 {fmt_hm(c.climb_h)}    {climb_dist_used:>5,.0f} nm   {fmt_kg(c.climb_kg)} kg")
    out.append(f"   Cruise                {fmt_hm(c.cruise_h)}    {cruise_dist:>5,.0f} nm   {fmt_kg(c.cruise_kg)} kg")
    out.append(f"   Descent               {fmt_hm(c.descent_h)}    {DESCENT_DIST_NM:>5,.0f} nm   {fmt_kg(c.descent_kg)} kg")
    out.append(f"   Taxi-in               {fmt_hm(c.taxi_in_h)}           —   {fmt_kg(c.taxi_in_kg)} kg")
    out.append(line)
    out.append(f"   Trip fuel                                     {fmt_kg(c.trip_kg)} kg")
    out.append(f"   BLOCK                 {fmt_hm(c.block_time_h)}                {fmt_kg(c.block_kg)} kg")
    out.append(line)
    out.append("  RESERVES")
    out.append(f"   Contingency  ({c.contingency_pct*100:.0f} % trip)                       {fmt_kg(c.contingency_kg)} kg")
    out.append(f"   Alternate    ({c.alternate_nm:.0f} nm)                      {fmt_kg(c.alternate_kg)} kg")
    out.append(f"   Final reserve ({c.final_reserve_min:.0f} min)                        {fmt_kg(c.final_reserve_kg)} kg")
    out.append(f"   Total reserves                                {fmt_kg(c.reserves_kg)} kg")
    out.append(dline)
    out.append(f"   TOTAL FUEL REQUIRED                           {fmt_kg(c.total_fuel_kg)} kg")
    pct_cap = c.total_fuel_kg / ac.max_fuel * 100 if ac.max_fuel else 0
    out.append(f"     tank capacity {fmt_kg(ac.max_fuel)} kg   "
               f"({pct_cap:5.1f} % used)")
    out.append(line)
    out.append("  WEIGHTS (estimate)")
    if c.zfw_kg is not None:
        out.append(f"   ZFW   {fmt_kg(c.zfw_kg)} kg    (MZFW {fmt_kg(ac.mzfw)} kg)")
        out.append(f"   TOW   {fmt_kg(c.tow_kg)} kg    (MTOW {fmt_kg(ac.mtow)} kg)")
        out.append(f"   LW    {fmt_kg(c.lw_kg)} kg    (MLW  {fmt_kg(ac.mlw)} kg)")
    else:
        out.append("   (OEW not published for this variant — weight checks skipped)")
    out.append(dline)
    if c.warnings:
        out.append("  WARNINGS")
        for w in c.warnings:
            out.append(f"   !  {w}")
        out.append(dline)
    out.append("  🦥  Route data: @ianolaf".rjust(L))
    return "\n".join(out)



# Interactive selection (I miss the interactive Netflix Series Minecraft Story Mode :( )

def _ask(prompt: str, default: Optional[str] = None) -> str:
    suffix = f" [{default}]" if default else ""
    raw = input(f"  {prompt}{suffix}: ").strip()
    return raw or (default or "")


def _ask_float(prompt: str, default: float) -> float:
    while True:
        raw = _ask(prompt, f"{default:g}")
        try:
            return float(raw)
        except ValueError:
            print("    (enter a number)")


def interactive(aircraft: Dict[str, Aircraft],
                routes: Dict[Tuple[str, str], float]) -> Calculation:
    print()
    print("═══════════════════════════════════════════════════════════")
    print("                            .--.  ")
    print("                           /    \\ ")
    print("                          | o  o |")
    print("                           \\  vv/ ")
    print("                           /`--'\\ ")
    print("                          (      )   🦥  takes its time,")
    print("                         /|      |\\      arrives on schedule")
    print("                        (_(______)_)")
    print(" TRANSATLANTIC FUEL & FLIGHT-TIME CALCULATOR")
    print("  Route data courtesy of @ianolaf")
    print("═══════════════════════════════════════════════════════════")

    origins = sorted({o for (o, _) in routes})
    print("\n Origins in route database:")
    for i, code in enumerate(origins, 1):
        end = "  " if i % 6 else "\n"
        print(f"   {code}", end=end)
    if len(origins) % 6:
        print()

    while True:
        origin = _ask("Origin ICAO").upper()
        if origin in origins:
            break
        print("    (not in database)")

    dests = sorted({d for (o, d) in routes if o == origin})
    print(f"\n Destinations reachable from {origin}:")
    for code in dests:
        print(f"   {code}  ({routes[(origin, code)]:,.0f} nm)")

    while True:
        dest = _ask("Destination ICAO").upper()
        if dest in dests:
            break
        print("    (not reachable from origin)")

    distance = routes[(origin, dest)]
    route = Route(origin=origin, destination=dest, distance_nm=distance)
    print(f"\n  → Distance {origin} – {dest}: {distance:,.0f} nm")

    # Aircraft selection (I wish I had the kind of money to select my aircraft in real life... )
    fams: Dict[str, List[Aircraft]] = {}
    for ac in aircraft.values():
        fams.setdefault(ac.family, []).append(ac)
    print("\n Aircraft (✓ = range OK for this route):")
    for fam in sorted(fams):
        print(f"   [{fam}]")
        for ac in sorted(fams[fam], key=lambda a: a.variant):
            ok = "✓" if ac.max_range >= distance else "·"
            print(f"     {ok} {ac.id:<13}  {ac.variant:<14}"
                  f"  range {ac.max_range:>5,.0f} nm   burn {ac.cruise_burn:>5,.0f} kg/h")
    while True:
        raw = _ask("Aircraft ID")
        key = next((k for k in aircraft if k.lower() == raw.lower()), None)
        if key:
            chosen = aircraft[key]
            break
        print("    (unknown ID)")

    print("\n Optional parameters (Enter = default):")
    default_payload = estimate_payload(chosen)
    payload       = _ask_float("Payload kg", round(default_payload))
    wind          = _ask_float("Wind component kt (+ headwind, - tailwind)", 0)
    final_reserve = _ask_float("Final-reserve holding time (min)", FINAL_RESERVE_MIN)
    alternate     = _ask_float("Alternate distance (nm)", ALTERNATE_DIST_NM)
    contingency   = _ask_float("Contingency (% of trip fuel)", CONTINGENCY_PCT * 100)

    c = calculate(
        chosen, route,
        payload_kg=payload,
        wind_component_kt=wind,
        alternate_nm=alternate,
        final_reserve_min=final_reserve,
        contingency_pct=contingency / 100.0,
    )
    print()
    print(format_report(c))
    return c


# List modes

def list_aircraft(aircraft: Dict[str, Aircraft]) -> None:
    print(f"{'ID':<13} {'Manufacturer':<11} {'Family':<9} {'Variant':<14} "
          f"{'Range':>7} {'Burn':>8} {'MaxFuel':>9}")
    print("─" * 80)
    for ac in sorted(aircraft.values(),
                     key=lambda a: (a.manufacturer, a.family, a.variant)):
        print(f"{ac.id:<13} {ac.manufacturer:<11} {ac.family:<9} "
              f"{ac.variant:<14} {ac.max_range:>6,.0f}  "
              f"{ac.cruise_burn:>6,.0f}  {ac.max_fuel:>9,.0f}")


def list_routes(routes: Dict[Tuple[str, str], float]) -> None:
    seen = set()
    for (o, d), dist in sorted(routes.items()):
        if (d, o) in seen:
            continue
        seen.add((o, d))
        print(f"   {o}  ↔  {d}    {dist:>6,.0f} nm")



# Main

def main(argv: Optional[List[str]] = None) -> int:
    p = argparse.ArgumentParser(
        description="Transatlantic fuel & flight-time calculator")
    p.add_argument("--aircraft-xlsx",
                   help="Path to aircraft workbook (auto-detected if omitted)")
    p.add_argument("--routes-xlsx",
                   help="Path to routes workbook (auto-detected if omitted)")
    p.add_argument("--origin",   help="Origin ICAO")
    p.add_argument("--dest",     help="Destination ICAO")
    p.add_argument("--aircraft", help="Aircraft ID from database")
    p.add_argument("--payload", type=float, help="Payload in kg")
    p.add_argument("--wind",    type=float, default=0.0,
                   help="Wind component kt (+ headwind)")
    p.add_argument("--taxi-out", type=float, default=TAXI_OUT_MIN_DEFAULT)
    p.add_argument("--taxi-in",  type=float, default=TAXI_IN_MIN_DEFAULT)
    p.add_argument("--alternate-nm", type=float, default=ALTERNATE_DIST_NM)
    p.add_argument("--final-reserve", type=float, default=FINAL_RESERVE_MIN,
                   help="Final-reserve holding time in minutes "
                        f"(default {FINAL_RESERVE_MIN:.0f})")
    p.add_argument("--contingency", type=float, default=CONTINGENCY_PCT * 100,
                   help="Contingency as %% of trip fuel "
                        f"(default {CONTINGENCY_PCT * 100:.0f})")
    p.add_argument("--list-aircraft", action="store_true")
    p.add_argument("--list-routes",   action="store_true")
    args = p.parse_args(argv)

    ac_path = Path(args.aircraft_xlsx) if args.aircraft_xlsx else _pick_existing(AIRCRAFT_CANDIDATES)
    rt_path = Path(args.routes_xlsx)   if args.routes_xlsx   else _pick_existing(ROUTES_CANDIDATES)
    if ac_path is None or not ac_path.exists():
        print("Aircraft workbook not found.", file=sys.stderr)
        return 2
    if rt_path is None or not rt_path.exists():
        print("Routes workbook not found.", file=sys.stderr)
        return 2

    aircraft = load_aircraft(ac_path)
    routes   = load_routes(rt_path)

    if args.list_aircraft:
        list_aircraft(aircraft)
        return 0
    if args.list_routes:
        list_routes(routes)
        return 0

    if args.origin and args.dest and args.aircraft:
        key = (args.origin.upper(), args.dest.upper())
        if key not in routes:
            print(f"Route {key[0]} → {key[1]} not in database.", file=sys.stderr)
            return 1
        ac = aircraft.get(args.aircraft) or \
             next((v for k, v in aircraft.items() if k.lower() == args.aircraft.lower()), None)
        if ac is None:
            print(f"Aircraft '{args.aircraft}' not in database.  "
                  "Try --list-aircraft.", file=sys.stderr)
            return 1
        route = Route(origin=key[0], destination=key[1], distance_nm=routes[key])
        c = calculate(ac, route,
                      payload_kg=args.payload, wind_component_kt=args.wind,
                      taxi_out_min=args.taxi_out, taxi_in_min=args.taxi_in,
                      alternate_nm=args.alternate_nm,
                      final_reserve_min=args.final_reserve,
                      contingency_pct=args.contingency / 100.0)
        print(format_report(c))
        return 0

    try:
        interactive(aircraft, routes)
    except (KeyboardInterrupt, EOFError):
        print("\nInterrupted.")
        return 130
    except Exception as e:
        import traceback
        print("\nUnexpected error:", e)
        traceback.print_exc()
        try:
            input("\nPress Enter to close this window...")
        except Exception:
            pass
        return 1

    # Keep the console open when the script was launched by double-clicking
    # on Windows — otherwise the window closes instantly after the report.
    try:
        input("\nPress Enter to close this window...")
    except Exception:
        pass
    return 0


if __name__ == "__main__":
    sys.exit(main())
     